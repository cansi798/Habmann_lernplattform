"""Excel-Strukturpläne → data/course-plan.json.

Liest beide Strukturplan-Excel-Dateien (Kurs 22 und Kurs 71) aus dem
übergeordneten Projektverzeichnis und erzeugt ein voll befülltes
course-plan.json mit 16 Tagen pro Kurs inklusive aller UE-Inhalte.

Excel-Layout (beide Dateien identisch):
  R1:  Spalten-Header (UE | Level | Methodik | Tag 1 ... Tag 16)
  R2:  "Thema des Tages" — Modul-Header + Schwerpunkt
  R3:  "Kompetenzentwicklung" — Lernziele L1/L2/L3
  R4:  UE 1 (Level 1 Wissen) — Vermittelte Inhalte
  R5:  UE 2 (Level 1 Arbeitsauftrag 1)
  R6:  UE 3 (Level 1 Arbeitsauftrag 2)
  R7:  UE 4 (Level 1+2 Coaching)
  R8:  UE 5 (Level 2 Anwendung) — Fallstudie
  R9:  UE 6 (oft leer)
  R10: UE 7 (Level 2+3 Feedback)
  R11: UE 8 (Senior Level 3 Transferprojekt)
  R12-13: UE 9, 10 (meist leer)
"""
from __future__ import annotations
import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTER_ROOT = REPO_ROOT.parent
DATA_DIR = REPO_ROOT / "data"

XLSX_FILES = [
    {
        "kurs_id": "22_Bo_143",
        "kurz": "Vertrieb & Plattformen",
        "titel": "Digitale Vertriebsstrategien & Plattform-Geschäftsmodelle",
        "akzentfarbe": "#1a4d8f",
        "akzent_bg": "#dceaff",
        "path": OUTER_ROOT / "[22_Bo_143] Digitale Vertriebsstrategien & Plattform-Geschäftsmodelle im Vertrieb entwickeln und steuern" / "Strukturplan 22_Bo_143 (05-26).xlsx",
    },
    {
        "kurs_id": "71_Dig_42",
        "kurz": "Business Modelling",
        "titel": "Digital Business Modelling & Transformation",
        "akzentfarbe": "#9d4a1a",
        "akzent_bg": "#ffe9d6",
        "path": OUTER_ROOT / "[71_Dig_42] Digital Business Modelling & Transformation – IT-gestützte Geschäftsmodelle und Veränderungsprozesse umsetzen" / "Strukturplan 71_Dig_42 (05-26) (3).xlsx",
    },
]

# Datumsplan: Mo-Fr, mit Fronleichnam 2026-06-04 als Pause.
# Ergibt 16 Unterrichtstage zwischen Do 2026-05-28 und Fr 2026-06-19.
KURSTAGE = [
    ("2026-05-28", "Do"), ("2026-05-29", "Fr"),
    ("2026-06-01", "Mo"), ("2026-06-02", "Di"), ("2026-06-03", "Mi"),
    # 04.06. Fronleichnam — frei
    ("2026-06-05", "Fr"),
    ("2026-06-08", "Mo"), ("2026-06-09", "Di"), ("2026-06-10", "Mi"),
    ("2026-06-11", "Do"), ("2026-06-12", "Fr"),
    ("2026-06-15", "Mo"), ("2026-06-16", "Di"), ("2026-06-17", "Mi"),
    ("2026-06-18", "Do"), ("2026-06-19", "Fr"),
]
assert len(KURSTAGE) == 16, f"Datumsplan muss 16 Tage haben, hat {len(KURSTAGE)}"


def parse_thema(cell: str) -> tuple[dict, str]:
    """R2 splitten in Modul-Metadaten + Schwerpunkt-Text.

    Beispiel-Input:
      'Modul 1 (Tag 1 von 2): Digitale Vertriebsstrategien verstehen ...
       Schwerpunkt des Tages ist ...'
    """
    if not cell:
        return ({}, "")
    # Header-Zeile bis zum ersten Doppel-Newline ist die Modul-Beschreibung
    parts = re.split(r"\n\n+", cell, maxsplit=1)
    header = parts[0].strip()
    body = parts[1].strip() if len(parts) > 1 else ""

    # Header parsen: "Modul N (Tag X von Y): Titel" — "von Y" und ":" optional
    m = re.match(r"Modul\s+(\d+)\s*\(Tag\s+(\d+)(?:\s+von\s+(\d+))?\)\s*:?\s*(.+)", header, re.DOTALL)
    if m:
        modul = {
            "nr": int(m.group(1)),
            "tag_in_modul": int(m.group(2)),
            "tage_gesamt": int(m.group(3)) if m.group(3) else int(m.group(2)),
            "titel": m.group(4).strip(),
        }
    else:
        modul = {"nr": 0, "tag_in_modul": 0, "tage_gesamt": 0, "titel": header}

    return modul, body


def parse_lernziele(cell: str) -> dict:
    """R3 splitten in L1/L2/L3.

    Beispiel-Input:
      'Level 1 – Wissen
       • Grundlagen ...
       • Marktanalyse ...

       Level 2 – Anwendung
       • ...'
    """
    if not cell:
        return {"wissen": [], "anwendung": [], "management": []}

    result = {"wissen": [], "anwendung": [], "management": []}
    current = None
    for line in cell.split("\n"):
        line = line.strip()
        if not line:
            continue
        low = line.lower()
        if "level 1" in low and ("wissen" in low or "verständnis" in low):
            current = "wissen"
        elif "level 2" in low and "anwendung" in low:
            current = "anwendung"
        elif "level 3" in low and ("management" in low or "transfer" in low):
            current = "management"
        elif line.startswith(("•", "·", "-", "*")):
            if current:
                result[current].append(line.lstrip("•·-* ").strip())
    return result


def extract_day(ws, col: int) -> dict:
    """Pro Tag (Spalte col) das komplette Daten-Dict aus 13 Zeilen."""
    thema_cell = ws.cell(2, col).value or ""
    lernziele_cell = ws.cell(3, col).value or ""
    modul, schwerpunkt = parse_thema(thema_cell)
    lernziele = parse_lernziele(lernziele_cell)

    ue_rows = []
    for r in range(4, 14):
        label = ws.cell(r, 1).value
        level = ws.cell(r, 2).value
        methodik = ws.cell(r, 3).value
        inhalt = ws.cell(r, col).value
        if not inhalt:
            continue
        ue_rows.append({
            "nr": int(label) if isinstance(label, (int, str)) and str(label).isdigit() else None,
            "level": (level or "").strip(),
            "methodik": (methodik or "").strip(),
            "inhalt": inhalt.strip(),
        })

    return {
        "modul": modul,
        "schwerpunkt": schwerpunkt,
        "lernziele": lernziele,
        "ue": ue_rows,
    }


def build_course(kurs_meta: dict) -> dict:
    wb = openpyxl.load_workbook(kurs_meta["path"], data_only=True)
    ws = wb["Tabelle1"]
    tage = []
    for tag_nr in range(1, 17):  # Tag 1-16 in Spalten 4-19
        col = 3 + tag_nr
        day_data = extract_day(ws, col)
        datum, wochentag = KURSTAGE[tag_nr - 1]
        modul = day_data["modul"]
        tag = {
            "nr": tag_nr,
            "datum": datum,
            "wochentag": wochentag,
            "modul": {"nr": modul.get("nr", 0), "titel": modul.get("titel", "")},
            "tag_im_modul": f"{modul.get('tag_in_modul', 0)}/{modul.get('tage_gesamt', 0)}",
            "thema": modul.get("titel", ""),
            "schwerpunkt": day_data["schwerpunkt"],
            "lernziele": day_data["lernziele"],
            "ue": day_data["ue"],
        }
        tage.append(tag)

    return {
        "id": kurs_meta["kurs_id"],
        "kurz": kurs_meta["kurz"],
        "titel": kurs_meta["titel"],
        "akzentfarbe": kurs_meta["akzentfarbe"],
        "akzent_bg": kurs_meta["akzent_bg"],
        "tage": tage,
    }


def main():
    kurse = [build_course(k) for k in XLSX_FILES]
    out = {"kurse": kurse}
    target = DATA_DIR / "course-plan.json"
    target.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    for k in kurse:
        ue_total = sum(len(t["ue"]) for t in k["tage"])
        print(f"  {k['id']}: 16 Tage, {ue_total} UEs gesamt")
    print(f"Geschrieben: {target.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
